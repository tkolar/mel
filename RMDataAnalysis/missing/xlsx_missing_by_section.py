import sys
import csv
import openpyxl

subsection_corrections = {
    'Judd Dolle Span': 'Judd Dolle Spanish',
    'Judd Dolle - Re': 'Judd Dolle - Reference',
    'Hardcover fict' : 'Hardcover fiction',
    '2nd language re' : '2nd language resource',
    'child inter rea' : 'child inter reader',
    'pic. books pape ' : 'pic. books paper',
    'child inter rea' : 'child inter reader',
    'child adv. read' : 'child adv. reader',
    'child mid reade' : 'child mid reader'
    }


section_map = {
    'Hardcover fict': 'Fiction',
    'Fiction': 'Fiction',
    'Short Stories': 'Fiction',
    'Fantasy': 'Fiction',
    'Horror': 'Fiction',
    'Science Fiction': 'Fiction',

    'Non-fiction': 'Non-fiction',
    '2nd language re': 'Non-fiction',

    'Judd Dolle': 'Judd Dolle',
    'Judd Dolle Span': 'Judd Dolle',
    'Judd Dolle - Re': 'Judd Dolle',

    'Audiobooks': 'Audiobooks',
    'DVD': 'DVD',
    'Puzzles': 'Puzzles',

    '': 'Unknown',
    'Holidays': 'Holidays',

    'Young Adult': 'Young Adult',

    'Children': 'Children',
    'child 900': 'Children',
    'pic. books hard': 'Children',
    'child inter rea': 'Children',
    'child 300': 'Children',
    'pic. books pape': 'Children',
    'child adv. read': 'Children',
    'child 500': 'Children',
    'child 600': 'Children',
    'child mid reade': 'Children',
    'child. Latin Am': 'Children',
    'child early rea': 'Children',
    'Child B.C.': 'Children',
    'large print': 'Children',
    'child 900': 'Children',
    'child anth.': 'Children',
    "Children's DVD": 'Children',
    "Children's Puzz": 'Children',
    'child 800': 'Children',
    'Board Books': 'Children'
    }

#
# items.csv fields
#
# 'Resource Type', 'Title', 'Section', 'Author', 'Publisher', 'Status', '# times Checked Out', 'Accession Date', 'Barcode'
#
def get_book_dict():
    book_dict = {}
    csvfile = open("items.csv", encoding="latin-1")
    cvsreader = csv.DictReader(csvfile)
    for row in cvsreader:
        barcode = row["Barcode"]
        book_dict[barcode] = row

    return(book_dict)


def main():

    book_dict = get_book_dict()

    from openpyxl import Workbook


    shelf_line = 0
    shelf_name = ""
    cubby = ""
    location_names = []
    locations = {}

    #
    #  Go through all the missing items
    #
    all = {}

    section_list = set()
    
    barfile = open(sys.argv[1])
    for line in barfile:
        sline = line[:-1]
        if sline in book_dict:
            section = book_dict[sline]["Section"]
            if section in subsection_corrections:
                book_dict[sline]["Section"] = subsection_corrections[section]

            section = section_map[section]
            section_list.add(section)
            if section not in all:
                all[section] = []
            all[section].append(sline)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Missing Books By Section"

    row = 1
    for main_section in sorted(section_list):
        ws1.cell(column = 1, row = row, value = "Section")
        ws1.cell(column = 2, row = row, value = "Resource Type")
        ws1.cell(column = 3, row = row, value = "Title")
        ws1.cell(column = 4, row = row, value = "Author")
        ws1.cell(column = 5, row = row, value = "# times Checked Out")
        ws1.cell(column = 6, row = row, value = "Barcode")
        row += 1
        for barcode in all[main_section]:
            if barcode in book_dict:
                ws1.cell(column = 1, row = row, value = book_dict[barcode]["Section"])
                ws1.cell(column = 2, row = row, value = book_dict[barcode]["Resource Type"])
                ws1.cell(column = 3, row = row, value = book_dict[barcode]["Title"])
                ws1.cell(column = 4, row = row, value = book_dict[barcode]["Author"])
                ws1.cell(column = 5, row = row, value = book_dict[barcode]["# times Checked Out"])
                ws1.cell(column = 6, row = row, value = book_dict[barcode]["Barcode"])
                row += 1
            else:
                print("Missing barcode: %s" % barcode)
        row += 2

    wb.save("missing_by_section.xlsx")

main()

